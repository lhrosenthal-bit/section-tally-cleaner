import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
from datetime import datetime

# Title
st.title("📊 Section Tally Cleaner")

# File uploader
uploaded_file = st.file_uploader("Upload your Section Tally Excel file", type=["xlsx"])

if uploaded_file:
    st.success("File uploaded successfully!")

    # Load workbook
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    # Extract semester info from first non-empty cell in row 7
    semester = None
    for cell in ws[7]:
        if cell.value:
            semester = str(cell.value).strip()
            break

    # Unmerge cells
    for merged_cell in list(ws.merged_cells):
        ws.unmerge_cells(str(merged_cell))

    # Remove empty columns
    for col in reversed(range(1, ws.max_column + 1)):
        if all([ws.cell(row=row, column=col).value in [None, ""] for row in range(1, ws.max_row + 1)]):
            ws.delete_cols(col)

    # Fill down 'Program' column
    program_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=8, column=col).value == "Program":
            program_col = col
            break
    if program_col:
        last_value = None
        for row in range(9, ws.max_row + 1):
            cell = ws.cell(row=row, column=program_col)
            if cell.value:
                last_value = cell.value
            else:
                cell.value = last_value

    # Remove embedded images
    ws._images.clear()

    # Remove first 7 rows
    ws.delete_rows(1, 7)

    # Insert semester info into column E and rename header
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).value = semester
    ws.cell(row=1, column=5).value = "Semester"

    # Freeze top row
    ws.freeze_panes = "A2"

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Save to BytesIO
    output = BytesIO()
    month_year = datetime.now().strftime("%B %Y")
    filename = f"Section Tally Final {semester} downloaded {month_year}.xlsx"
    wb.save(output)
    output.seek(0)

    # Download button
    st.download_button(
        label="📥 Download Cleaned File",
        data=output,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
